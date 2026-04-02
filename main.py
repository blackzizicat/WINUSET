import json
import os
import csv
import re
import warnings
from datetime import date, timedelta

import openpyxl
import requests

warnings.filterwarnings('ignore')

API_BASE_URL = (
    'https://script.google.com/macros/s/'
    'AKfycbyG6q-YCMLN4ppMOZpa-ZRH1SvVpKshuH0eFaCgfb2CD9lqVaZ6J4MtI0Ehpe5sqgf0/exec'
)
CONFIG_PATH = '/credentials/config.json'


def get_prev_month():
    """実行日の前月を (year, month) で返す"""
    today = date.today()
    first = today.replace(day=1)
    prev = first - timedelta(days=1)
    return prev.year, prev.month


def parse_csv(filepath):
    """shift_jis の CSV を読み込み、行のリストで返す"""
    with open(filepath, encoding='shift_jis') as f:
        reader = csv.reader(f)
        return list(reader)


def to_value(s):
    """文字列を int / float / str / None に変換する"""
    if s == '':
        return None
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return s


def month_to_fiscal_index(month):
    """4月始まり年度における月→インデックス（3=4月, 4=5月, ..., 11=12月, 12=1月, 13=2月, 14=3月）"""
    if month >= 4:
        return month - 1
    else:
        return month + 11



def get_api_key():
    """credentials/config.json からシークレットキーを読み込む"""
    if not os.path.exists(CONFIG_PATH):
        raise FileNotFoundError(
            f'設定ファイルが見つかりません: {CONFIG_PATH}\n'
            '{"api_key": "YOUR_SECRET_KEY"} の形式で作成してください。'
        )
    with open(CONFIG_PATH) as f:
        return json.load(f)['api_key']


def fetch_winclient_count():
    """メール取得 API から「Windowsクライアント月次統計」の利用回数を取得する"""
    api_key = get_api_key()
    url = f'{API_BASE_URL}?type=winclient&key={api_key}'
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    body = resp.json()['body']

    for line in body.splitlines():
        if '■Windowsクライアント利用回数:' in line:
            # 数値部分を抽出（カンマ区切りにも対応）
            m = re.search(r'[\d,]+', line.split(':', 1)[1])
            if m:
                return int(m.group().replace(',', ''))

    raise ValueError(f'「■Windowsクライアント利用回数:」が見つかりませんでした。\nbody:\n{body[:500]}')


def fetch_spss_ccmaster_count():
    """SPSS使用者数月次レポートから ccmasterドメインの Statictics 利用回数を取得する"""
    api_key = get_api_key()
    url = f'{API_BASE_URL}?type=spss&key={api_key}'
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    body = resp.json()['body']

    in_section = False
    for line in body.splitlines():
        if '■Statictics利用回数' in line:
            in_section = True
            continue
        if in_section and line.strip().startswith('■'):
            break  # 次のセクションに入ったら終了
        if in_section and 'ccmasterドメイン（教卓PC，共用PC等）' in line:
            m = re.search(r'[\d,]+', line.split('：', 1)[1])
            if m:
                return int(m.group().replace(',', ''))

    raise ValueError(f'「ccmasterドメイン（教卓PC，共用PC等）」が見つかりませんでした。\nbody:\n{body[:500]}')


def fetch_onedrive_data():
    """OneDrive月次ログインレポートから3種の利用者数を取得する。
    戻り値: (必ず使用, 使用するときもある, 使用しない)
    """
    api_key = get_api_key()
    url = f'{API_BASE_URL}?type=onedrive&key={api_key}'
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    body = resp.json()['body']

    def extract(label):
        for line in body.splitlines():
            if label in line:
                m = re.search(r'[\d,]+', line.split('：', 1)[1])
                if m:
                    return int(m.group().replace(',', ''))
        raise ValueError(f'「{label}」が見つかりませんでした。')

    always   = extract('OneDriveを必ず使用')
    sometime = extract('OneDriveを使用するときもある')
    never    = extract('OneDriveを使用しない')
    return always, sometime, never


def fetch_winapp_data():
    """メール取得 API から「Windowsアプリケーション月次統計」の利用回数を取得する。
    戻り値: {アプリ名(小文字): 利用回数} の辞書
    """
    api_key = get_api_key()
    url = f'{API_BASE_URL}?type=winapp&key={api_key}'
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    body = resp.json()['body']

    app_counts = {}
    in_section = False
    for line in body.splitlines():
        if '■Windowsアプリケーション利用回数' in line:
            in_section = True
            continue
        if not in_section:
            continue
        parts = line.split(',')
        if len(parts) == 4:
            app_name = parts[1].strip()
            try:
                count = int(parts[3].strip().replace(',', ''))
            except ValueError:
                continue
            app_counts[app_name.lower()] = count

    if not app_counts:
        raise ValueError(f'アプリ利用回数データが取得できませんでした。\nbody:\n{body[:500]}')
    return app_counts


def trend_word(curr, prev, threshold=50, is_final=False, zero_means_none=False):
    """増減判定ワードを返す。
    差が threshold 以内 → 横ばい、超過 → 増加、以下 → 減少。
    zero_means_none=True かつ curr==0 のとき → 「なし」。
    is_final=True のとき語尾に「した/だった」を付ける。
    """
    if zero_means_none and (curr is None or curr == 0):
        return 'なし'
    if curr is None or prev is None:
        return '不明'
    diff = curr - prev
    if diff > threshold:
        return '増加した' if is_final else '増加し'
    elif diff < -threshold:
        return '減少した' if is_final else '減少し'
    else:
        return '横ばいだった' if is_final else '横ばい'


def get_user_count_from_sheet(wb, sheet_name):
    """月次シートから統計PC利用者数（学生+院生+教員+職員+misc+対象外）を計算する。
    Winログインシート AC列の数式と同じ集計ロジック。
    """
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    try:
        student  = sum(ws.cell(row=r, column=13).value or 0 for r in [18, 19, 20, 21])
        grad     = ws.cell(row=29, column=4).value or 0
        faculty  = ws.cell(row=36, column=4).value or 0
        staff    = ws.cell(row=43, column=1).value or 0
        misc     = ws.cell(row=50, column=1).value or 0
        excluded = ws.cell(row=53, column=1).value or 0
        return student + grad + faculty + staff + misc + excluded
    except Exception:
        return None


def update_user_count(report_dir, month):
    """月次シートから利用者数を計算し、Winログインシートの AC列に直接書き込む。
    AC列は本来 INDIRECT 数式で計算されるが、openpyxl では数式を評価できないため
    数値を直接入力する運用とする。
    """
    excel_path = os.path.join(report_dir, '07_Windows利用統計.xlsx')
    if not os.path.exists(excel_path):
        print(f'[SKIP] Excel が見つかりません: 07_Windows利用統計.xlsx')
        return

    sheet_name = f'{month:02d}'
    wb = openpyxl.load_workbook(excel_path)
    count = get_user_count_from_sheet(wb, sheet_name)
    if count is None:
        print(f'  [WARN] 利用者数の計算に失敗しました（シート: {sheet_name}）')
        return

    row = month_to_fiscal_index(month)
    ws = wb['Winログイン']
    ws.cell(row=row, column=29).value = count  # AC列 = 列29
    wb.save(excel_path)
    print(f'利用者数を書き込みました: {count}名  (Winログイン!AC{row})')
    print()


def get_top_apps(wb, month, n=6):
    """Winアプリシートから当月の利用回数上位 n 件のアプリ名を返す。"""
    ws = wb['Winアプリ']
    col = month_to_fiscal_index(month)  # 3=4月 … 14=3月（列番号と一致）
    app_values = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        app_name = row[1]          # B列
        if not app_name:
            continue
        value = row[col - 1] if len(row) >= col else None  # 0-indexed
        if isinstance(value, (int, float)) and value > 0:
            app_values.append((app_name, value))
    app_values.sort(key=lambda x: x[1], reverse=True)
    return [name for name, _ in app_values[:n]]


def gen_report_text(report_dir, year, month):
    """レポート文章を生成して返す。"""
    prev_month = month - 1 if month > 1 else 12

    excel_win = os.path.join(report_dir, '07_Windows利用統計.xlsx')
    excel_od  = os.path.join(report_dir, '月次OneDriveログイン状況.xlsx')

    wb_win = openpyxl.load_workbook(excel_win, data_only=True)
    wb_od  = openpyxl.load_workbook(excel_od,  data_only=True)

    # 利用者数・利用回数（前月→今月）
    row_curr = month_to_fiscal_index(month)
    row_prev = month_to_fiscal_index(prev_month)
    ws_win = wb_win['Winログイン']
    curr_user  = ws_win.cell(row=row_curr, column=29).value  # AC列
    prev_user  = ws_win.cell(row=row_prev, column=29).value
    curr_count = ws_win.cell(row=row_curr, column=32).value  # AF列
    prev_count = ws_win.cell(row=row_prev, column=32).value

    user_word  = trend_word(curr_user,  prev_user,  is_final=False)
    count_word = trend_word(curr_count, prev_count, is_final=True)

    def fmt_num(v):
        """None を '-' に、数値をカンマ区切りで返す"""
        return '-' if v is None else f'{v:,}'

    # アプリ上位 6 件
    top_apps  = get_top_apps(wb_win, month)
    apps_text = '，'.join(top_apps)

    # OneDrive ログイン状況（前月→今月）
    od_row_curr = month_to_fiscal_index(month) - 1
    od_row_prev = month_to_fiscal_index(prev_month) - 1
    ws_od = wb_od['OneDriveログイン']
    od_login_curr = ws_od.cell(row=od_row_curr, column=2).value or 0  # B=利用者層
    od_login_prev = ws_od.cell(row=od_row_prev, column=2).value or 0
    od_both_curr  = ws_od.cell(row=od_row_curr, column=3).value or 0  # C=中間層
    od_both_prev  = ws_od.cell(row=od_row_prev, column=3).value or 0
    od_never_curr = ws_od.cell(row=od_row_curr, column=4).value or 0  # D=非利用者層
    od_never_prev = ws_od.cell(row=od_row_prev, column=4).value or 0

    login_word = trend_word(od_login_curr, od_login_prev, is_final=False, zero_means_none=True)
    both_word  = trend_word(od_both_curr,  od_both_prev,  is_final=False)
    never_word = trend_word(od_never_curr, od_never_prev, is_final=True)

    return (
        f'{year}年{month}月 利用統計レポート\n'
        f'渡邉さん\n'
        f'新庄です。\n'
        f'{month}月分の利用統計レポートを作成しましたのでご確認ください。\n'
        f'----------\n'
        f'・Windowsクライアント\n'
        f'利用者数は{user_word}，利用回数は{count_word}'
        f'（{fmt_num(prev_user)}名→{fmt_num(curr_user)}名，{fmt_num(prev_count)}回→{fmt_num(curr_count)}回）。'
        f'主に利用されているアプリケーションは{apps_text}だった。'
        f'教研AD配下のWindowsクライアントでのOneDriveログイン状況は，'
        f'利用者層は{login_word}，中間層は{both_word}，非利用者層は{never_word}。\n'
        f'----------\n'
        f'よろしくお願いいたします。'
    )


def write_csv_to_sheet(ws, rows):
    """既存シートの値をクリアし、CSV の内容で上書きする（書式は保持）"""
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.value = None
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c).value = to_value(val)


def main():
    year, month = get_prev_month()
    yyyymm = f"{year}{month:02d}"
    sheet_name = f"{month:02d}"

    source_dir = '/source'
    report_dir = '/report'

    mappings = [
        ('07_Windows利用統計.xlsx',    f'集計_winlogin_user_{yyyymm}.csv'),
        ('OneDriveログインしない.xlsx', f'集計_unuseOneDrive_user_{yyyymm}.csv'),
        ('OneDriveログインする.xlsx',   f'集計_useOneDrive_user_{yyyymm}.csv'),
        ('OneDriveログイン両方.xlsx',   f'集計_dualOneDrive_user_{yyyymm}.csv'),
    ]

    print(f'対象年月: {yyyymm}  書き込み先シート: [{sheet_name}]')
    print()

    # APIからデータを取得して各シートに書き込む
    update_winclient_count(report_dir, month)
    update_winapp_sheet(report_dir, month)
    update_onedrive_sheet(report_dir, month)


    for excel_name, csv_name in mappings:
        csv_path = os.path.join(source_dir, csv_name)
        excel_path = os.path.join(report_dir, excel_name)

        if not os.path.exists(csv_path):
            print(f'[SKIP] CSV が見つかりません: {csv_name}')
            continue
        if not os.path.exists(excel_path):
            print(f'[SKIP] Excel が見つかりません: {excel_name}')
            continue

        print(f'処理中: {csv_name}')
        print(f'  -> {excel_name} [sheet: {sheet_name}]')

        rows = parse_csv(csv_path)
        wb = openpyxl.load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            print(f'  [WARN] シート "{sheet_name}" が存在しません。')
            print(f'         利用可能シート: {wb.sheetnames}')
            continue

        ws = wb[sheet_name]
        write_csv_to_sheet(ws, rows)
        wb.save(excel_path)
        print(f'  -> 完了: {len(rows)} 行を書き込みました')
        print()

    # 月次シートから利用者数を計算して Winログイン!AC列に書き込む
    update_user_count(report_dir, month)

    # レポート文章を生成して出力
    print(gen_report_text(report_dir, year, month))


def update_winapp_sheet(report_dir, month):
    """Winアプリシートのアプリ利用回数を API から取得した値で更新する"""
    excel_path = os.path.join(report_dir, '07_Windows利用統計.xlsx')
    if not os.path.exists(excel_path):
        print(f'[SKIP] Excel が見つかりません: 07_Windows利用統計.xlsx')
        return

    print('APIからWindowsアプリケーション利用回数を取得中...')
    try:
        app_counts = fetch_winapp_data()
    except Exception as e:
        print(f'  [ERROR] winapp API取得失敗: {e}')
        return

    # SPSSはSPSS使用者数レポートの ccmasterドメイン値で上書き
    try:
        spss_count = fetch_spss_ccmaster_count()
        app_counts['spss'] = spss_count
        print(f'  -> SPSS(ccmaster): {spss_count}')
    except Exception as e:
        print(f'  [WARN] SPSS API取得失敗（winappの値を使用）: {e}')

    col = month_to_fiscal_index(month)  # C=3(4月) 〜 N=14(3月)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Winアプリ']

    written, skipped = 0, 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        app_name_cell = row[1]  # B列
        if app_name_cell.value is None:
            continue
        key = str(app_name_cell.value).lower()
        if key in app_counts:
            ws.cell(row=app_name_cell.row, column=col).value = app_counts[key]
            written += 1
        else:
            skipped += 1

    wb.save(excel_path)
    print(f'  -> 完了: {written} 行を書き込み（{skipped} 行はスキップ）  列: {col} ({month}月)')
    print()


def update_onedrive_sheet(report_dir, month):
    """OneDriveログインシートのB,C,D列をAPIから取得した値で更新する"""
    excel_path = os.path.join(report_dir, '月次OneDriveログイン状況.xlsx')
    if not os.path.exists(excel_path):
        print(f'[SKIP] Excel が見つかりません: 月次OneDriveログイン状況.xlsx')
        return

    print('APIからOneDriveログイン状況を取得中...')
    try:
        always, sometime, never = fetch_onedrive_data()
    except Exception as e:
        print(f'  [ERROR] API取得失敗: {e}')
        return

    row = month_to_fiscal_index(month) - 1  # 4月=row2, 2月=row12, 3月=row13
    print(f'  -> 必ず使用:{always}  両方:{sometime}  使用しない:{never}  行:{row}')

    wb = openpyxl.load_workbook(excel_path)
    ws = wb['OneDriveログイン']
    ws.cell(row=row, column=2).value = always    # B列
    ws.cell(row=row, column=3).value = sometime  # C列
    ws.cell(row=row, column=4).value = never     # D列
    wb.save(excel_path)
    print(f'  -> 完了')
    print()


def update_winclient_count(report_dir, month):
    """Winログインシートの AF 列（利用回数）を API から取得した値で更新する"""
    excel_path = os.path.join(report_dir, '07_Windows利用統計.xlsx')
    if not os.path.exists(excel_path):
        print(f'[SKIP] Excel が見つかりません: 07_Windows利用統計.xlsx')
        return

    print('APIからWindowsクライアント利用回数を取得中...')
    try:
        count = fetch_winclient_count()
    except Exception as e:
        print(f'  [ERROR] API取得失敗: {e}')
        return

    row = month_to_fiscal_index(month)
    print(f'  -> 取得値: {count}  書き込み先: Winログイン!AF{row}')

    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Winログイン']
    ws.cell(row=row, column=32).value = count  # 列32 = AF
    wb.save(excel_path)
    print(f'  -> 完了')
    print()


if __name__ == '__main__':
    main()
